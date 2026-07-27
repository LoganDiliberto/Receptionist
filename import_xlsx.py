"""One-shot importer: read the legacy salon workbook and populate the DB.

Usage:
    python -m import_xlsx <path/to/ReceptionistData.xlsx>
    python import_xlsx.py <path/to/ReceptionistData.xlsx>

Idempotent-ish safety: refuses to run if the target DB already contains
services, staff, or appointments unless ``--force`` is passed. Pass
``--wipe`` to truncate the salon tables before importing (this is what
``entrypoint.sh`` uses when it first sees an xlsx on the Fly volume).

Reads the same six sheets ``salon.py`` used to load from directly:
Hours, Location, Services, Associates, Schedule, Appointments.
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
from loguru import logger
from sqlalchemy import delete, select

from db import session_scope
from models import (
    Appointment,
    Client as ClientRow,
    Hours as HoursRow,
    Salon as SalonRow,
    Service as ServiceRow,
    Staff as StaffRow,
    StaffHours as StaffHoursRow,
)


def _normalize_phone(raw: str | None) -> str:
    """Match salon.normalize_phone. Kept local to avoid the salon import
    from firing (it eagerly loads INFO, which is wasted work during import)."""
    if not raw:
        return ""
    digits = re.sub(r"\D", "", raw)
    return digits[-10:] if len(digits) >= 10 else ""


WEEKDAYS = ["Sunday", "Monday", "Tuesday", "Wednesday",
            "Thursday", "Friday", "Saturday"]

_APPT_COLUMNS = [
    "id", "created_at", "customer_name", "customer_phone",
    "stylist", "service", "date", "start_time", "end_time",
    "session_id",
]


# ---------- Parsers (mirror salon.py, kept here so this script has no
#            circular import back into the running app). ----------


def _parse_time(s: str) -> time:
    s = s.strip().upper().replace(" ", "")
    m = re.match(r"(\d{1,2})(?::(\d{2}))?(AM|PM)", s)
    if not m:
        raise ValueError(f"Cannot parse time: {s!r}")
    hour = int(m.group(1))
    minute = int(m.group(2) or 0)
    if m.group(3) == "PM" and hour < 12:
        hour += 12
    if m.group(3) == "AM" and hour == 12:
        hour = 0
    return time(hour, minute)


def _parse_clock(s: str) -> time:
    s = s.strip().upper()
    if re.search(r"AM|PM", s):
        return _parse_time(s)
    h, m = s.split(":")
    return time(int(h), int(m))


def _parse_range(s: str) -> tuple[time, time]:
    parts = re.split(r"\s+to\s+", s.strip(), flags=re.IGNORECASE)
    return _parse_time(parts[0]), _parse_time(parts[1])


def _coerce_time(v) -> time:
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    return _parse_clock(str(v))


def _coerce_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v), "%Y-%m-%d").date()


# ---------- Workbook parsing ----------


def _load_hours(wb) -> dict[str, tuple[time, time] | None]:
    out: dict[str, tuple[time, time] | None] = {}
    if "Hours" not in wb.sheetnames:
        return out
    for row in list(wb["Hours"].iter_rows(values_only=True))[1:]:
        day, open_s, close_s = row[0], row[1], row[2]
        if not day:
            continue
        if isinstance(open_s, str) and open_s.strip().upper() == "NA":
            out[str(day).strip()] = None
        elif open_s is None or close_s is None:
            out[str(day).strip()] = None
        else:
            out[str(day).strip()] = (_parse_time(str(open_s)), _parse_time(str(close_s)))
    return out


def _load_location(wb) -> str:
    if "Location" not in wb.sheetnames:
        return ""
    rows = list(wb["Location"].iter_rows(values_only=True))
    return str(rows[1][1]) if len(rows) > 1 and rows[1][1] else ""


def _load_services(wb) -> list[tuple[str, int, float]]:
    if "Services" not in wb.sheetnames:
        return []
    out: list[tuple[str, int, float]] = []
    for row in list(wb["Services"].iter_rows(values_only=True))[1:]:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        try:
            duration = int(row[1]) if row[1] is not None else 30
        except (TypeError, ValueError):
            duration = 30
        try:
            price = float(row[2]) if row[2] is not None else 0.0
        except (TypeError, ValueError):
            price = 0.0
        out.append((name, duration, price))
    return out


def _load_staff(wb) -> dict[str, tuple[list[str], dict[str, tuple[time, time]]]]:
    """Return {staff_name: (services_lc, {weekday: (start, end)})}."""
    associates: dict[str, list[str]] = {}
    if "Associates" in wb.sheetnames:
        for row in list(wb["Associates"].iter_rows(values_only=True))[1:]:
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            services = [
                str(s).strip().lower()
                for s in row[1:]
                if s and str(s).strip()
            ]
            associates[name] = services

    schedules: dict[str, dict[str, tuple[time, time]]] = {}
    if "Schedule" in wb.sheetnames:
        rows = list(wb["Schedule"].iter_rows(values_only=True))
        if rows:
            day_headers = [d for d in rows[0][1:] if d]
            for row in rows[1:]:
                if not row or not row[0]:
                    continue
                name = str(row[0]).strip()
                sched: dict[str, tuple[time, time]] = {}
                for day, cell in zip(day_headers, row[1:]):
                    if cell and isinstance(cell, str) and cell.strip():
                        sched[str(day).strip()] = _parse_range(cell)
                schedules[name] = sched

    all_names = set(associates) | set(schedules)
    return {n: (associates.get(n, []), schedules.get(n, {})) for n in all_names}


def _load_appointments(wb) -> list[dict]:
    if "Appointments" not in wb.sheetnames:
        return []
    ws = wb["Appointments"]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if not row or row[0] is None:
            continue
        rec = dict(zip(header, row))
        out.append(rec)
    return out


def _load_clients(wb) -> list[dict]:
    """Parse the (Phase-2 addition) Clients sheet. Missing sheet is fine —
    older workbooks pre-date the clients feature; they just import
    zero clients."""
    if "Clients" not in wb.sheetnames:
        return []
    rows = list(wb["Clients"].iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    out: list[dict] = []
    for row in rows[1:]:
        if not row or all(c is None for c in row):
            continue
        rec = dict(zip(header, row))
        # phone is required — skip rows without one so we don't hit the
        # NOT NULL constraint mid-import.
        if not rec.get("phone"):
            continue
        out.append(rec)
    return out


# ---------- Table-level ops ----------


def _has_data(sess) -> bool:
    for model in (ServiceRow, StaffRow, Appointment):
        if sess.scalar(select(model).limit(1)) is not None:
            return True
    return False


def _wipe(sess) -> None:
    """Truncate every salon table. FK order matters — junction + child rows first."""
    sess.execute(delete(Appointment))
    sess.execute(delete(StaffHoursRow))
    # staff_services is a join table; issue raw SQL since we don't have a
    # model class for it in scope. (The StaffService model exists but is
    # marked as `viewonly=False` on the relationship, meaning it's the
    # relationship's job to sync — we bypass it here for a truncate.)
    from models import StaffService
    sess.execute(delete(StaffService))
    sess.execute(delete(StaffRow))
    sess.execute(delete(ServiceRow))
    sess.execute(delete(HoursRow))
    sess.execute(delete(SalonRow))
    sess.execute(delete(ClientRow))


# ---------- Main import ----------


def import_workbook(path: Path, *, wipe: bool = False, force: bool = False) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    logger.info(f"Reading workbook: {path}")

    location = _load_location(wb)
    hours = _load_hours(wb)
    services = _load_services(wb)
    staff = _load_staff(wb)
    appointments = _load_appointments(wb)
    clients = _load_clients(wb)

    logger.info(
        f"Parsed: location={bool(location)}, hours={sum(1 for h in hours.values() if h)}"
        f", services={len(services)}, staff={len(staff)}"
        f", clients={len(clients)}, appointments={len(appointments)}"
    )

    with session_scope() as sess:
        if _has_data(sess) and not (wipe or force):
            raise SystemExit(
                "Refusing to import: DB already contains salon data. "
                "Re-run with --wipe (truncates first) or --force (merges)."
            )
        if wipe:
            logger.info("Wiping existing salon tables before import.")
            _wipe(sess)
            sess.flush()

        # Location — single row, id=1.
        existing = sess.scalar(select(SalonRow).order_by(SalonRow.id).limit(1))
        if existing is None:
            sess.add(SalonRow(id=1, location=location))
        else:
            existing.location = location

        # Hours — one row per weekday.
        for day in WEEKDAYS:
            slot = hours.get(day)
            row = sess.scalar(select(HoursRow).where(HoursRow.weekday == day))
            if row is None:
                row = HoursRow(weekday=day)
                sess.add(row)
            if slot is None:
                row.open = None
                row.close = None
            else:
                row.open, row.close = slot

        # Services — upsert by name.
        service_by_name_lc: dict[str, ServiceRow] = {}
        for name, duration, price in services:
            existing_svc = sess.scalar(select(ServiceRow).where(ServiceRow.name.ilike(name)))
            if existing_svc is None:
                existing_svc = ServiceRow(name=name, duration_minutes=duration, price=price)
                sess.add(existing_svc)
            else:
                existing_svc.duration_minutes = duration
                existing_svc.price = price
            service_by_name_lc[name.lower()] = existing_svc
        sess.flush()

        # Staff — upsert, then rebuild services + hours.
        for staff_name, (offered_lc, sched) in staff.items():
            row = sess.scalar(select(StaffRow).where(StaffRow.name.ilike(staff_name)))
            if row is None:
                row = StaffRow(name=staff_name)
                sess.add(row)
                sess.flush()

            # Services
            row.services = []
            sess.flush()
            for svc_lc in offered_lc:
                svc = service_by_name_lc.get(svc_lc)
                if svc is None:
                    logger.warning(
                        f"Staff {staff_name!r} references unknown service {svc_lc!r} — skipping."
                    )
                    continue
                row.services.append(svc)

            # Schedule
            for h in list(row.hours):
                sess.delete(h)
            sess.flush()
            for day, (start, end) in sched.items():
                row.hours.append(StaffHoursRow(weekday=day, start=start, end=end))
        sess.flush()

        # Clients — upsert by normalized phone. Skipping any legacy rows
        # missing phone happens earlier in _load_clients.
        client_by_phone: dict[str, ClientRow] = {}
        for rec in clients:
            phone = _normalize_phone(str(rec.get("phone") or ""))
            if not phone:
                logger.warning(
                    f"Skipping client row with unusable phone {rec.get('phone')!r}"
                )
                continue
            existing_client = sess.scalar(select(ClientRow).where(ClientRow.phone == phone))
            first = str(rec.get("first_name") or "").strip()
            last = str(rec.get("last_name") or "").strip()
            email = (str(rec["email"]).strip() or None) if rec.get("email") else None
            gender = (str(rec["gender"]).strip() or None) if rec.get("gender") else None
            notes = (str(rec["notes"]).strip() or None) if rec.get("notes") else None

            def _dt(raw) -> datetime:
                if isinstance(raw, datetime):
                    return raw
                if raw:
                    try:
                        return datetime.fromisoformat(str(raw))
                    except ValueError:
                        pass
                return datetime.now()

            created_at = _dt(rec.get("created_at"))
            updated_at = _dt(rec.get("updated_at")) if rec.get("updated_at") else created_at

            if existing_client is None:
                existing_client = ClientRow(
                    first_name=first,
                    last_name=last,
                    phone=phone,
                    email=email,
                    gender=gender,
                    notes=notes,
                    created_at=created_at,
                    updated_at=updated_at,
                )
                sess.add(existing_client)
            else:
                existing_client.first_name = first or existing_client.first_name
                existing_client.last_name = last or existing_client.last_name
                existing_client.email = email or existing_client.email
                existing_client.gender = gender or existing_client.gender
                existing_client.notes = notes or existing_client.notes
                existing_client.updated_at = updated_at
            client_by_phone[phone] = existing_client
        sess.flush()

        # Make sure any existing client rows (from --force) are indexed too,
        # so appointments can still link if the xlsx only carries phone
        # numbers on Appointment rows.
        for row in sess.scalars(select(ClientRow)):
            client_by_phone.setdefault(row.phone, row)

        # Appointments — insert by original id, skip any we already have
        # (matters if the caller passed --force on a non-empty DB).
        staff_by_name_lc = {
            r.name.lower(): r
            for r in sess.scalars(select(StaffRow))
        }
        for rec in appointments:
            appt_id = rec.get("id")
            if not appt_id:
                continue
            appt_id = str(appt_id).strip()
            if sess.get(Appointment, appt_id) is not None:
                continue

            try:
                created_raw = rec.get("created_at")
                if isinstance(created_raw, datetime):
                    created_at = created_raw
                elif created_raw:
                    created_at = datetime.fromisoformat(str(created_raw))
                else:
                    created_at = datetime.now()

                stylist_name = str(rec.get("stylist") or "").strip()
                service_name = str(rec.get("service") or "").strip()
                staff_row = staff_by_name_lc.get(stylist_name.lower())
                svc_row = service_by_name_lc.get(service_name.lower())
                if not staff_row or not svc_row:
                    logger.warning(
                        f"Skipping appointment {appt_id!r}: "
                        f"unknown stylist={stylist_name!r} or service={service_name!r}"
                    )
                    continue

                # Link this appointment to a Client row if we can find one
                # by phone. This is the retroactive-linking hook: legacy
                # workbooks without a Clients sheet still hydrate the FK
                # from the phone column that's always been there.
                customer_phone_raw = str(rec.get("customer_phone") or "").strip()
                normalized = _normalize_phone(customer_phone_raw)
                client_row = client_by_phone.get(normalized) if normalized else None

                appt = Appointment(
                    id=appt_id,
                    created_at=created_at,
                    customer_name=str(rec.get("customer_name") or "").strip(),
                    customer_phone=customer_phone_raw,
                    staff_id=staff_row.id,
                    service_id=svc_row.id,
                    client_id=client_row.id if client_row else None,
                    date=_coerce_date(rec.get("date")),
                    start_time=_coerce_time(rec.get("start_time")),
                    end_time=_coerce_time(rec.get("end_time")),
                    session_id=(str(rec["session_id"]).strip()
                                if rec.get("session_id") else None),
                )
                sess.add(appt)
            except Exception as e:
                logger.warning(f"Skipping appointment {appt_id!r}: {e}")

    logger.success(f"Import complete: {path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import a legacy salon workbook into the SQLite DB."
    )
    parser.add_argument("path", type=Path, help="Path to ReceptionistData.xlsx")
    parser.add_argument(
        "--wipe", action="store_true",
        help="Truncate salon tables before importing (safest for a fresh migration).",
    )
    parser.add_argument(
        "--force", action="store_true",
        help="Import even if the DB already has data (merges by primary key).",
    )
    args = parser.parse_args()
    try:
        import_workbook(args.path, wipe=args.wipe, force=args.force)
    except SystemExit as e:
        print(str(e), file=sys.stderr)
        sys.exit(2)


if __name__ == "__main__":
    main()
