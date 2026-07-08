"""One-off: dump the structure of the salon Excel sheet so we can design tools."""
from pathlib import Path
import openpyxl

path = Path(r"C:\Users\logan\Downloads\Receptionist data.xlsx")
wb = openpyxl.load_workbook(path, data_only=True)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print("=" * 70)
    print(f"Sheet: {sheet_name!r}   dims={ws.dimensions}   rows={ws.max_row}  cols={ws.max_column}")
    print("=" * 70)
    for row in ws.iter_rows(values_only=True):
        # Skip fully-empty rows
        if all(v is None for v in row):
            continue
        print(row)
    print()
