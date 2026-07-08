"""Salon data store: hours, staff, services, schedules, and appointments.

Backed by an Excel workbook. On startup, ensures every sheet the app expects
exists (creating a default `Services` sheet from historical constants if the
workbook doesn't have one yet). Static info is cached in `INFO`; mutations go
through the CRUD helpers below, which persist to the workbook and refresh the
cache so the next voice-bot call sees the new state.

Public surface used by the voice bot (`bot.py`):
  - INFO                       (cached SalonInfo)
  - SERVICE_DURATIONS_MIN      (derived {name.lower(): minutes})
  - system_prompt_context()    (rendered block for the system prompt)
  - check_availability(...)    (LLM tool)
  - book_appointment(...)      (LLM tool)

Public surface used by the admin API (`server.py`):
  - list/create/update/delete for staff, services, appointments
  - get/update for hours and location
  - reload() to refresh in-memory state after a mutation
"""

from __future__ import annotations

import asyncio
import os
import re
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from threading import RLock
from typing import Iterable

import openpyxl
from loguru import logger

DEFAULT_XLSX = Path(__file__).parent / "ReceptionistData.xlsx"
SALON_XLSX = Path(os.getenv("SALON_DATA_PATH", str(DEFAULT_XLSX)))

# The canonical order we render weekdays in everywhere (schedules, hours, UI).
WEEKDAYS = ["Sunday", "Monday", "Tuesday", "Wednesday",
            "Thursday", "Friday", "Saturday"]

# Fallback used the very first time we boot against an older workbook that
# doesn't yet have a Services sheet. Real prices are trivially editable in
# the admin UI once the sheet exists.
_DEFAULT_SERVICES: list[tuple[str, int, float]] = [
    ("Cut", 30, 35.0),
    ("Color", 90, 95.0),
    ("Perm", 120, 120.0),
    ("Shave", 30, 25.0),
]

SLOT_MIN = 30  # all appointments align to 30-minute slot boundaries

# Excel writes aren't atomic — serialize all workbook ops behind a single lock
# so concurrent admin edits and bot bookings don't stomp each other. We use a
# reentrant threading lock (not asyncio) because helpers call each other and
# are dispatched to `asyncio.to_thread`.
_lock = RLock()
_async_lock = asyncio.Lock()


# ---------- Parsing helpers ----------


def _parse_time(s: str) -> time:
    """Parse '10AM', '12 PM', '7PM', '10:00 AM' -> datetime.time."""
    s = s.strip().upper().replace(" ", "")
    m = re.match(r"(\d{1,2})(?::(\d{2}))?(AM|PM)", s)
    if not m:
        raise ValueError(f"Cannot parse time: {s!r}")
    hour = int(m.group(1))
    minute = int(m.group(2) or 0)
    if m.group(3) == "PM" and hour < 12:
        hour += 12
    if m.group(3) == "AM" and hour == 12:
        hour = 0
    return time(hour, minute)


def _parse_range(s: str) -> tuple[time, time]:
    """Parse '10AM to 4PM' -> (time(10,0), time(16,0))."""
    parts = re.split(r"\s+to\s+", s.strip(), flags=re.IGNORECASE)
    return _parse_time(parts[0]), _parse_time(parts[1])


def _parse_date(s: str) -> date:
    return datetime.strptime(s, "%Y-%m-%d").date()


def _parse_clock(s: str) -> time:
    """Parse '14:30' or '2:30 PM' -> datetime.time."""
    s = s.strip().upper()
    if re.search(r"AM|PM", s):
        return _parse_time(s)
    h, m = s.split(":")
    return time(int(h), int(m))


def _weekday_name(d: date) -> str:
    # Python's Monday=0, Sunday=6.
    return ["Monday", "Tuesday", "Wednesday", "Thursday",
            "Friday", "Saturday", "Sunday"][d.weekday()]


def _fmt_range(start: time, end: time) -> str:
    """Render a schedule range in the '10AM to 4PM' style the sheet uses."""

    def one(t: time) -> str:
        suffix = "AM" if t.hour < 12 else "PM"
        h = t.hour % 12 or 12
        return f"{h}:{t.minute:02d}{suffix}" if t.minute else f"{h}{suffix}"

    return f"{one(start)} to {one(end)}"


# ---------- Data model ----------


@dataclass(frozen=True)
class Service:
    name: str
    duration_minutes: int
    price: float


@dataclass(frozen=True)
class Stylist:
    name: str
    services: tuple[str, ...]  # lowercased service names
    schedule: dict[str, tuple[time, time]]  # weekday -> (start, end)


@dataclass
class SalonInfo:
    location: str
    hours: dict[str, tuple[time, time] | None]  # weekday -> (open, close) or None
    stylists: dict[str, Stylist]
    services: dict[str, Service] = field(default_factory=dict)


# ---------- Workbook bootstrap ----------


_APPT_COLUMNS = [
    "id", "created_at", "customer_name", "customer_phone",
    "stylist", "service", "date", "start_time", "end_time",
]


def _ensure_workbook_shape() -> None:
    """Make sure every sheet we depend on exists, seeding sane defaults."""
    with _lock:
        wb = openpyxl.load_workbook(SALON_XLSX)
        dirty = False

        if "Appointments" not in wb.sheetnames:
            ws = wb.create_sheet("Appointments")
            ws.append(_APPT_COLUMNS)
            dirty = True

        if "Services" not in wb.sheetnames:
            ws = wb.create_sheet("Services")
            ws.append(["name", "duration_minutes", "price"])
            for name, dur, price in _DEFAULT_SERVICES:
                ws.append([name, dur, price])
            dirty = True

        if dirty:
            wb.save(SALON_XLSX)


# ---------- Read-side: load a fresh SalonInfo ----------


def _load_info() -> SalonInfo:
    wb = openpyxl.load_workbook(SALON_XLSX, data_only=True)

    # Hours sheet
    hours: dict[str, tuple[time, time] | None] = {}
    for day, open_s, close_s in list(wb["Hours"].iter_rows(values_only=True))[1:]:
        if not day:
            continue
        if (open_s or "").strip().upper() == "NA":
            hours[day] = None
        else:
            hours[day] = (_parse_time(str(open_s)), _parse_time(str(close_s)))

    # Location sheet
    loc_rows = list(wb["Location"].iter_rows(values_only=True))
    location = str(loc_rows[1][1]) if len(loc_rows) > 1 and loc_rows[1][1] else ""

    # Services sheet
    services: dict[str, Service] = {}
    for row in list(wb["Services"].iter_rows(values_only=True))[1:]:
        if not row or not row[0]:
            continue
        name = str(row[0]).strip()
        try:
            duration = int(row[1]) if row[1] is not None else 30
        except (TypeError, ValueError):
            duration = 30
        try:
            price = float(row[2]) if row[2] is not None else 0.0
        except (TypeError, ValueError):
            price = 0.0
        services[name.lower()] = Service(name=name, duration_minutes=duration, price=price)

    # Associates sheet: staff -> services they offer
    stylist_services: dict[str, tuple[str, ...]] = {}
    for row in list(wb["Associates"].iter_rows(values_only=True))[1:]:
        name = row[0]
        if not name:
            continue
        offered = tuple(
            s.strip().lower()
            for s in row[1:]
            if s and isinstance(s, str) and s.strip()
        )
        stylist_services[str(name).strip()] = offered

    # Schedule sheet: staff -> weekday -> (start, end)
    sched_rows = list(wb["Schedule"].iter_rows(values_only=True))
    day_headers = [d for d in sched_rows[0][1:] if d]
    schedules: dict[str, dict[str, tuple[time, time]]] = {}
    for row in sched_rows[1:]:
        name = row[0]
        if not name:
            continue
        sched: dict[str, tuple[time, time]] = {}
        for day, cell in zip(day_headers, row[1:]):
            if cell and isinstance(cell, str) and cell.strip():
                sched[day] = _parse_range(cell)
        schedules[str(name).strip()] = sched

    stylists = {
        name: Stylist(name=name, services=offered, schedule=schedules.get(name, {}))
        for name, offered in stylist_services.items()
    }
    return SalonInfo(location=location, hours=hours, stylists=stylists, services=services)


_ensure_workbook_shape()
INFO: SalonInfo = _load_info()
SERVICE_DURATIONS_MIN: dict[str, int] = {
    key: svc.duration_minutes for key, svc in INFO.services.items()
}
logger.info(
    f"Salon loaded from {SALON_XLSX.name}: "
    f"{len(INFO.stylists)} staff ({', '.join(INFO.stylists) or 'none'}), "
    f"{len(INFO.services)} services"
)


def reload() -> None:
    """Refresh cached state after a workbook mutation. Cheap; loads ~5 sheets."""
    global INFO, SERVICE_DURATIONS_MIN
    with _lock:
        INFO = _load_info()
        SERVICE_DURATIONS_MIN = {
            key: svc.duration_minutes for key, svc in INFO.services.items()
        }


# ---------- Appointments sheet R/W ----------


def _read_appointments() -> list[dict]:
    with _lock:
        wb = openpyxl.load_workbook(SALON_XLSX, data_only=True)
        ws = wb["Appointments"]
        rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    out = []
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        rec = dict(zip(_APPT_COLUMNS, row))
        # Normalise: dates/times may come back as datetime objects if Excel
        # auto-typed the cell. We always want ISO strings on the wire.
        if isinstance(rec.get("date"), (datetime, date)):
            rec["date"] = rec["date"].strftime("%Y-%m-%d") if not isinstance(rec["date"], datetime) else rec["date"].date().isoformat()
        else:
            rec["date"] = str(rec["date"])
        rec["start_time"] = _normalize_clock_str(rec.get("start_time"))
        rec["end_time"] = _normalize_clock_str(rec.get("end_time"))
        out.append(rec)
    return out


def _normalize_clock_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.time().strftime("%H:%M")
    return str(v)


def _append_appointment(appt: dict) -> None:
    with _lock:
        wb = openpyxl.load_workbook(SALON_XLSX)
        ws = wb["Appointments"]
        ws.append([appt[k] for k in _APPT_COLUMNS])
        wb.save(SALON_XLSX)


def _rewrite_appointments(appts: list[dict]) -> None:
    """Replace the Appointments sheet contents wholesale. Called for edits/deletes."""
    with _lock:
        wb = openpyxl.load_workbook(SALON_XLSX)
        # Drop-and-recreate is the simplest way to reset without leaving stale rows.
        del wb["Appointments"]
        ws = wb.create_sheet("Appointments")
        ws.append(_APPT_COLUMNS)
        for appt in appts:
            ws.append([appt.get(k) for k in _APPT_COLUMNS])
        wb.save(SALON_XLSX)


# ---------- Public salon-context helper for the system prompt ----------


def _fmt_hour(t: time) -> str:
    """Cross-platform hour-only 12-hour format ('7 PM')."""
    suffix = "AM" if t.hour < 12 else "PM"
    h = t.hour % 12 or 12
    return f"{h} {suffix}"


def system_prompt_context() -> str:
    """Render salon static info as a block to inject into the LLM system prompt."""
    lines = ["SALON LOCATION:", f"  {INFO.location}", "", "SALON HOURS:"]
    for day in WEEKDAYS[1:] + [WEEKDAYS[0]]:  # Monday..Sunday
        h = INFO.hours.get(day)
        if h is None:
            lines.append(f"  {day}: closed")
        else:
            lines.append(f"  {day}: {_fmt_hour(h[0])} to {_fmt_hour(h[1])}")

    lines += ["", "STAFF AND SERVICES:"]
    for name, s in INFO.stylists.items():
        services = ", ".join(sorted(s.services))
        lines.append(f"  {name}: {services}")

    lines += ["", "STAFF SCHEDULES:"]
    for name, s in INFO.stylists.items():
        days = []
        for day in WEEKDAYS[1:] + [WEEKDAYS[0]]:
            if day in s.schedule:
                start, end = s.schedule[day]
                days.append(f"{day[:3]} {_fmt_range(start, end)}")
        lines.append(f"  {name}: {', '.join(days) if days else 'not scheduled'}")

    lines += ["", "SERVICE DURATIONS (minutes):"]
    for svc in INFO.services.values():
        lines.append(f"  {svc.name}: {svc.duration_minutes}")

    return "\n".join(lines)


# ---------- Slot math ----------


def _slots_in_range(start: time, end: time) -> list[time]:
    out = []
    cur = datetime.combine(date.today(), start)
    last = datetime.combine(date.today(), end)
    while cur + timedelta(minutes=SLOT_MIN) <= last:
        out.append(cur.time())
        cur += timedelta(minutes=SLOT_MIN)
    return out


def _overlaps(a_start: time, a_end: time, b_start: time, b_end: time) -> bool:
    return not (a_end <= b_start or a_start >= b_end)


# ---------- Tool: check_availability ----------


async def check_availability(
    date_iso: str,
    stylist: str | None = None,
    service: str | None = None,
) -> dict:
    """Return open appointment slots on the given date, optionally filtered."""
    async with _async_lock:
        return await asyncio.to_thread(_check_availability_sync, date_iso, stylist, service)


def _check_availability_sync(date_iso: str, stylist: str | None, service: str | None) -> dict:
    try:
        d = _parse_date(date_iso)
    except ValueError:
        return {"error": f"Invalid date {date_iso!r}. Use YYYY-MM-DD."}

    weekday = _weekday_name(d)

    if INFO.hours.get(weekday) is None:
        return {"date": date_iso, "weekday": weekday, "closed": True, "available": []}
    salon_open, salon_close = INFO.hours[weekday]

    candidates = list(INFO.stylists.values())
    if stylist:
        candidates = [s for s in candidates if s.name.lower() == stylist.strip().lower()]
        if not candidates:
            return {"error": f"Unknown stylist {stylist!r}. Available: {list(INFO.stylists)}."}

    service_lc = service.strip().lower() if service else None
    if service_lc:
        if service_lc not in SERVICE_DURATIONS_MIN:
            return {"error": f"Unknown service {service!r}. Offered: {list(SERVICE_DURATIONS_MIN)}."}
        candidates = [s for s in candidates if service_lc in s.services]
        if not candidates:
            return {"error": f"No stylist on the requested filter offers {service}."}

    duration = SERVICE_DURATIONS_MIN.get(service_lc, 30)

    booked = _read_appointments()
    by_stylist: dict[str, list[tuple[time, time]]] = {}
    for a in booked:
        if str(a["date"]) != date_iso:
            continue
        by_stylist.setdefault(a["stylist"], []).append(
            (_parse_clock(str(a["start_time"])), _parse_clock(str(a["end_time"])))
        )

    result = []
    for s in candidates:
        if weekday not in s.schedule:
            continue
        work_start, work_end = s.schedule[weekday]
        win_start = max(work_start, salon_open)
        win_end = min(work_end, salon_close)
        all_slots = _slots_in_range(win_start, win_end)
        their = by_stylist.get(s.name, [])
        avail: list[str] = []
        for start in all_slots:
            end_dt = datetime.combine(d, start) + timedelta(minutes=duration)
            end = end_dt.time()
            if end > win_end:
                continue
            if any(_overlaps(start, end, bs, be) for bs, be in their):
                continue
            avail.append(start.strftime("%H:%M"))
        if avail:
            result.append({"stylist": s.name, "times": avail})

    return {
        "date": date_iso,
        "weekday": weekday,
        "closed": False,
        "service": service_lc,
        "duration_minutes": duration,
        "available": result,
    }


# ---------- Tool: book_appointment ----------


async def book_appointment(
    customer_name: str,
    customer_phone: str,
    stylist: str,
    service: str,
    date_iso: str,
    time_str: str,
) -> dict:
    """Book and persist an appointment after re-checking availability."""
    async with _async_lock:
        return await asyncio.to_thread(
            _book_appointment_sync,
            customer_name, customer_phone, stylist, service, date_iso, time_str,
        )


def _book_appointment_sync(
    customer_name: str,
    customer_phone: str,
    stylist: str,
    service: str,
    date_iso: str,
    time_str: str,
) -> dict:
    try:
        d = _parse_date(date_iso)
        start = _parse_clock(time_str)
    except ValueError as e:
        return {"error": str(e)}

    stylist_obj = next(
        (s for s in INFO.stylists.values() if s.name.lower() == stylist.strip().lower()),
        None,
    )
    if not stylist_obj:
        return {"error": f"Unknown stylist {stylist!r}."}

    service_lc = service.strip().lower()
    if service_lc not in SERVICE_DURATIONS_MIN:
        return {"error": f"Unknown service {service!r}."}
    if service_lc not in stylist_obj.services:
        return {
            "error": f"{stylist_obj.name} does not offer {service}. "
                     f"They offer: {', '.join(sorted(stylist_obj.services))}."
        }

    weekday = _weekday_name(d)
    if INFO.hours.get(weekday) is None:
        return {"error": f"Salon is closed on {weekday}s."}
    salon_open, salon_close = INFO.hours[weekday]

    duration = SERVICE_DURATIONS_MIN[service_lc]
    end_dt = datetime.combine(d, start) + timedelta(minutes=duration)
    end = end_dt.time()

    if weekday not in stylist_obj.schedule:
        return {"error": f"{stylist_obj.name} doesn't work on {weekday}s."}
    work_start, work_end = stylist_obj.schedule[weekday]
    if start < work_start or end > work_end:
        return {
            "error": f"{stylist_obj.name} works "
                     f"{work_start.strftime('%I:%M %p')} - {work_end.strftime('%I:%M %p')} "
                     f"on {weekday}s. {duration}-minute appointment doesn't fit."
        }
    if start < salon_open or end > salon_close:
        return {"error": "Time is outside salon open hours."}

    for a in _read_appointments():
        if str(a["date"]) != date_iso or a["stylist"] != stylist_obj.name:
            continue
        ostart = _parse_clock(str(a["start_time"]))
        oend = _parse_clock(str(a["end_time"]))
        if _overlaps(start, end, ostart, oend):
            return {"error": f"{stylist_obj.name} is already booked at {time_str}."}

    appt_id = uuid.uuid4().hex[:8]
    _append_appointment({
        "id": appt_id,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "customer_name": customer_name,
        "customer_phone": customer_phone,
        "stylist": stylist_obj.name,
        "service": service_lc.title(),
        "date": date_iso,
        "start_time": start.strftime("%H:%M"),
        "end_time": end.strftime("%H:%M"),
    })

    return {
        "ok": True,
        "appointment_id": appt_id,
        "summary": (
            f"{service_lc.title()} with {stylist_obj.name} on {weekday} "
            f"{date_iso} at {start.strftime('%I:%M %p').lstrip('0')} ({duration} minutes)."
        ),
    }


# ==================== Admin CRUD API ====================
#
# Everything below is synchronous. The API layer runs each one in a thread
# and calls `reload()` after any mutation.


# ---------- Staff ----------


def list_staff() -> list[dict]:
    """Return all staff as JSON-friendly dicts."""
    out = []
    for name, s in INFO.stylists.items():
        out.append({
            "name": name,
            "services": [svc for svc in s.services],
            "schedule": {
                day: {
                    "start": start.strftime("%H:%M"),
                    "end": end.strftime("%H:%M"),
                }
                for day, (start, end) in s.schedule.items()
            },
        })
    return out


def _write_staff_sheets(staff: list[dict]) -> None:
    """Rewrite the Associates and Schedule sheets from a normalized staff list."""
    with _lock:
        wb = openpyxl.load_workbook(SALON_XLSX)

        # Associates: one row per person, up to 4 service columns (matches
        # the existing shape). We compute the width from the data so the
        # sheet grows if we ever exceed 4.
        if "Associates" in wb.sheetnames:
            del wb["Associates"]
        ws = wb.create_sheet("Associates")
        max_services = max((len(s.get("services", [])) for s in staff), default=0)
        max_services = max(max_services, 4)
        header = ["Name"] + [f"Service {i+1}" for i in range(max_services)]
        ws.append(header)
        for s in staff:
            services = list(s.get("services", []))
            row = [s["name"]] + [
                (services[i].title() if i < len(services) else None)
                for i in range(max_services)
            ]
            ws.append(row)

        # Schedule: columns are weekdays, rows are people, cells are ranges.
        if "Schedule" in wb.sheetnames:
            del wb["Schedule"]
        ws = wb.create_sheet("Schedule")
        ws.append([None] + WEEKDAYS)
        for s in staff:
            row = [s["name"]]
            for day in WEEKDAYS:
                slot = s.get("schedule", {}).get(day)
                if slot and slot.get("start") and slot.get("end"):
                    start = _parse_clock(slot["start"])
                    end = _parse_clock(slot["end"])
                    row.append(_fmt_range(start, end))
                else:
                    row.append(None)
            ws.append(row)

        wb.save(SALON_XLSX)


def _validate_staff_payload(s: dict) -> None:
    if not isinstance(s.get("name"), str) or not s["name"].strip():
        raise ValueError("Staff name is required.")
    if not isinstance(s.get("services", []), list):
        raise ValueError("services must be a list.")
    schedule = s.get("schedule", {})
    if not isinstance(schedule, dict):
        raise ValueError("schedule must be an object keyed by weekday.")
    for day, slot in schedule.items():
        if day not in WEEKDAYS:
            raise ValueError(f"Unknown weekday {day!r}. Use one of {WEEKDAYS}.")
        if slot is None:
            continue
        if not isinstance(slot, dict):
            raise ValueError(f"Schedule for {day} must be an object with start/end.")
        start = slot.get("start")
        end = slot.get("end")
        if not start and not end:
            continue
        try:
            st = _parse_clock(str(start))
            en = _parse_clock(str(end))
        except Exception as e:
            raise ValueError(f"Invalid time in {day} schedule: {e}") from e
        if st >= en:
            raise ValueError(f"{day} start must be before end.")


def _normalize_staff(s: dict) -> dict:
    """Normalize the payload we persist: lowercase service names, drop empties."""
    services = [str(x).strip().lower() for x in s.get("services", []) if str(x).strip()]
    schedule: dict[str, dict[str, str]] = {}
    for day, slot in (s.get("schedule") or {}).items():
        if not slot:
            continue
        start = slot.get("start")
        end = slot.get("end")
        if not start or not end:
            continue
        schedule[day] = {"start": start, "end": end}
    return {"name": s["name"].strip(), "services": services, "schedule": schedule}


def create_staff(payload: dict) -> dict:
    _validate_staff_payload(payload)
    payload = _normalize_staff(payload)
    with _lock:
        current = list_staff()
        if any(m["name"].lower() == payload["name"].lower() for m in current):
            raise ValueError(f"Staff member {payload['name']!r} already exists.")
        current.append(payload)
        _write_staff_sheets(current)
    reload()
    return payload


def update_staff(name: str, payload: dict) -> dict:
    _validate_staff_payload(payload)
    payload = _normalize_staff(payload)
    with _lock:
        current = list_staff()
        idx = next((i for i, m in enumerate(current) if m["name"].lower() == name.lower()), None)
        if idx is None:
            raise KeyError(f"Staff member {name!r} not found.")
        # If the caller renamed the person, make sure the new name isn't taken.
        if payload["name"].lower() != name.lower() and any(
            m["name"].lower() == payload["name"].lower() for m in current
        ):
            raise ValueError(f"Staff member {payload['name']!r} already exists.")
        current[idx] = payload
        _write_staff_sheets(current)
    reload()
    return payload


def delete_staff(name: str) -> None:
    with _lock:
        current = list_staff()
        remaining = [m for m in current if m["name"].lower() != name.lower()]
        if len(remaining) == len(current):
            raise KeyError(f"Staff member {name!r} not found.")
        _write_staff_sheets(remaining)
    reload()


# ---------- Services ----------


def list_services() -> list[dict]:
    return [
        {"name": svc.name, "duration_minutes": svc.duration_minutes, "price": svc.price}
        for svc in INFO.services.values()
    ]


def _write_services_sheet(services: Iterable[dict]) -> None:
    with _lock:
        wb = openpyxl.load_workbook(SALON_XLSX)
        if "Services" in wb.sheetnames:
            del wb["Services"]
        ws = wb.create_sheet("Services")
        ws.append(["name", "duration_minutes", "price"])
        for svc in services:
            ws.append([svc["name"], int(svc["duration_minutes"]), float(svc["price"])])
        wb.save(SALON_XLSX)


def _validate_service_payload(s: dict) -> None:
    if not isinstance(s.get("name"), str) or not s["name"].strip():
        raise ValueError("Service name is required.")
    try:
        duration = int(s.get("duration_minutes"))
    except (TypeError, ValueError) as e:
        raise ValueError("duration_minutes must be an integer.") from e
    if duration <= 0 or duration % SLOT_MIN != 0:
        raise ValueError(
            f"duration_minutes must be a positive multiple of {SLOT_MIN}."
        )
    try:
        price = float(s.get("price"))
    except (TypeError, ValueError) as e:
        raise ValueError("price must be a number.") from e
    if price < 0:
        raise ValueError("price must be non-negative.")


def _normalize_service(s: dict) -> dict:
    return {
        "name": s["name"].strip().title(),
        "duration_minutes": int(s["duration_minutes"]),
        "price": float(s["price"]),
    }


def create_service(payload: dict) -> dict:
    _validate_service_payload(payload)
    payload = _normalize_service(payload)
    with _lock:
        current = list_services()
        if any(svc["name"].lower() == payload["name"].lower() for svc in current):
            raise ValueError(f"Service {payload['name']!r} already exists.")
        current.append(payload)
        _write_services_sheet(current)
    reload()
    return payload


def update_service(name: str, payload: dict) -> dict:
    _validate_service_payload(payload)
    payload = _normalize_service(payload)
    with _lock:
        current = list_services()
        idx = next(
            (i for i, svc in enumerate(current) if svc["name"].lower() == name.lower()),
            None,
        )
        if idx is None:
            raise KeyError(f"Service {name!r} not found.")
        if payload["name"].lower() != name.lower() and any(
            svc["name"].lower() == payload["name"].lower() for svc in current
        ):
            raise ValueError(f"Service {payload['name']!r} already exists.")
        current[idx] = payload
        _write_services_sheet(current)
    reload()
    return payload


def delete_service(name: str) -> None:
    with _lock:
        current = list_services()
        remaining = [svc for svc in current if svc["name"].lower() != name.lower()]
        if len(remaining) == len(current):
            raise KeyError(f"Service {name!r} not found.")
        _write_services_sheet(remaining)
    reload()


# ---------- Hours & location ----------


def get_hours() -> dict[str, dict | None]:
    """Weekday -> {'open': 'HH:MM', 'close': 'HH:MM'} or None when closed."""
    out: dict[str, dict | None] = {}
    for day in WEEKDAYS:
        h = INFO.hours.get(day)
        if h is None:
            out[day] = None
        else:
            out[day] = {"open": h[0].strftime("%H:%M"), "close": h[1].strftime("%H:%M")}
    return out


def _fmt_hour_sheet(t: time) -> str:
    """Match the existing '10AM'/'7PM' format the Hours sheet already uses."""
    suffix = "AM" if t.hour < 12 else "PM"
    h = t.hour % 12 or 12
    return f"{h}{suffix}" if t.minute == 0 else f"{h}:{t.minute:02d}{suffix}"


def update_hours(payload: dict[str, dict | None]) -> dict:
    """Overwrite the Hours sheet from a {day: {open, close} | None} dict."""
    parsed: dict[str, tuple[time, time] | None] = {}
    for day in WEEKDAYS:
        slot = payload.get(day) if isinstance(payload, dict) else None
        if slot is None:
            parsed[day] = None
            continue
        try:
            o = _parse_clock(str(slot["open"]))
            c = _parse_clock(str(slot["close"]))
        except Exception as e:
            raise ValueError(f"Invalid hours for {day}: {e}") from e
        if o >= c:
            raise ValueError(f"{day}: open time must be before close time.")
        parsed[day] = (o, c)

    with _lock:
        wb = openpyxl.load_workbook(SALON_XLSX)
        if "Hours" in wb.sheetnames:
            del wb["Hours"]
        ws = wb.create_sheet("Hours")
        ws.append(["Day", "Open", "Close"])
        for day in WEEKDAYS:
            slot = parsed[day]
            if slot is None:
                ws.append([day, "NA", "NA"])
            else:
                ws.append([day, _fmt_hour_sheet(slot[0]), _fmt_hour_sheet(slot[1])])
        wb.save(SALON_XLSX)
    reload()
    return get_hours()


def get_location() -> str:
    return INFO.location


def update_location(location: str) -> str:
    if not isinstance(location, str):
        raise ValueError("location must be a string.")
    with _lock:
        wb = openpyxl.load_workbook(SALON_XLSX)
        if "Location" in wb.sheetnames:
            del wb["Location"]
        ws = wb.create_sheet("Location")
        ws.append(["Locations", None])
        ws.append(["Location 1", location.strip()])
        wb.save(SALON_XLSX)
    reload()
    return get_location()


# ---------- Appointments (admin) ----------


def list_appointments(start: str | None = None, end: str | None = None) -> list[dict]:
    """Return appointments in [start, end] (inclusive). Both are optional."""
    start_d = _parse_date(start) if start else None
    end_d = _parse_date(end) if end else None
    out = []
    for a in _read_appointments():
        try:
            d = _parse_date(str(a["date"]))
        except ValueError:
            continue
        if start_d and d < start_d:
            continue
        if end_d and d > end_d:
            continue
        out.append(a)
    # Sort so the UI doesn't have to.
    out.sort(key=lambda a: (a["date"], a["start_time"]))
    return out


def _validate_appointment_payload(a: dict, *, require_id: bool = False) -> None:
    if require_id and not a.get("id"):
        raise ValueError("id is required.")
    for k in ("customer_name", "customer_phone", "stylist", "service", "date", "start_time"):
        if not a.get(k):
            raise ValueError(f"{k} is required.")
    try:
        _parse_date(str(a["date"]))
    except ValueError as e:
        raise ValueError(f"Invalid date: {e}") from e
    try:
        _parse_clock(str(a["start_time"]))
    except ValueError as e:
        raise ValueError(f"Invalid start_time: {e}") from e


def _end_from_service(service: str, start_time: str) -> str:
    service_lc = service.strip().lower()
    duration = SERVICE_DURATIONS_MIN.get(service_lc)
    if duration is None:
        raise ValueError(f"Unknown service {service!r}.")
    start = _parse_clock(start_time)
    end_dt = datetime.combine(date.today(), start) + timedelta(minutes=duration)
    return end_dt.time().strftime("%H:%M")


def create_appointment(payload: dict) -> dict:
    _validate_appointment_payload(payload)
    with _lock:
        existing = _read_appointments()
        appt = {
            "id": uuid.uuid4().hex[:8],
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "customer_name": payload["customer_name"].strip(),
            "customer_phone": payload["customer_phone"].strip(),
            "stylist": payload["stylist"].strip(),
            "service": payload["service"].strip().title(),
            "date": str(payload["date"]),
            "start_time": _parse_clock(payload["start_time"]).strftime("%H:%M"),
            "end_time": payload.get("end_time") or _end_from_service(
                payload["service"], payload["start_time"]
            ),
        }
        _check_admin_conflict(appt, existing)
        _append_appointment(appt)
    return appt


def update_appointment(appt_id: str, payload: dict) -> dict:
    _validate_appointment_payload(payload)
    with _lock:
        existing = _read_appointments()
        idx = next((i for i, a in enumerate(existing) if a["id"] == appt_id), None)
        if idx is None:
            raise KeyError(f"Appointment {appt_id!r} not found.")
        appt = {
            **existing[idx],
            "customer_name": payload["customer_name"].strip(),
            "customer_phone": payload["customer_phone"].strip(),
            "stylist": payload["stylist"].strip(),
            "service": payload["service"].strip().title(),
            "date": str(payload["date"]),
            "start_time": _parse_clock(payload["start_time"]).strftime("%H:%M"),
        }
        appt["end_time"] = payload.get("end_time") or _end_from_service(
            payload["service"], payload["start_time"]
        )
        others = [a for i, a in enumerate(existing) if i != idx]
        _check_admin_conflict(appt, others)
        others.insert(idx, appt)
        _rewrite_appointments(others)
    return appt


def delete_appointment(appt_id: str) -> None:
    with _lock:
        existing = _read_appointments()
        remaining = [a for a in existing if a["id"] != appt_id]
        if len(remaining) == len(existing):
            raise KeyError(f"Appointment {appt_id!r} not found.")
        _rewrite_appointments(remaining)


def _check_admin_conflict(new: dict, existing: list[dict]) -> None:
    """Block obvious double-bookings when the admin schedules by hand."""
    start = _parse_clock(new["start_time"])
    end = _parse_clock(new["end_time"])
    for a in existing:
        if a["date"] != new["date"] or a["stylist"] != new["stylist"]:
            continue
        ostart = _parse_clock(a["start_time"])
        oend = _parse_clock(a["end_time"])
        if _overlaps(start, end, ostart, oend):
            raise ValueError(
                f"{new['stylist']} is already booked at "
                f"{a['start_time']}-{a['end_time']} on {new['date']}."
            )
