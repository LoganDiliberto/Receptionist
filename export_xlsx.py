"""Dump the salon DB back into a workbook with the same shape ``salon.py``
used to read from. Useful for:

  - Ad-hoc backups (``fly ssh console`` + ``python -m export_xlsx /tmp/backup.xlsx``).
  - Round-tripping data between local dev and prod for debugging.
  - Handing a salon owner a spreadsheet they can eyeball or archive.

Sheet layout matches the historic ``ReceptionistData.xlsx``:
  Location, Hours, Services, Associates, Schedule, Appointments.
"""

from __future__ import annotations

import argparse
from datetime import time
from pathlib import Path

import openpyxl
from loguru import logger
from sqlalchemy import select
from sqlalchemy.orm import selectinload

from db import session_scope
from models import (
    Appointment,
    Client as ClientRow,
    Hours as HoursRow,
    Salon as SalonRow,
    Service as ServiceRow,
    Staff as StaffRow,
)


WEEKDAYS = ["Sunday", "Monday", "Tuesday", "Wednesday",
            "Thursday", "Friday", "Saturday"]


def _fmt_hour_sheet(t: time) -> str:
    """Match the historic '10AM'/'7PM' format the Hours sheet uses."""
    suffix = "AM" if t.hour < 12 else "PM"
    h = t.hour % 12 or 12
    return f"{h}{suffix}" if t.minute == 0 else f"{h}:{t.minute:02d}{suffix}"


def _fmt_range(start: time, end: time) -> str:
    """'10AM to 4PM' — the shape the Schedule sheet has always used."""

    def one(t: time) -> str:
        suffix = "AM" if t.hour < 12 else "PM"
        h = t.hour % 12 or 12
        return f"{h}:{t.minute:02d}{suffix}" if t.minute else f"{h}{suffix}"

    return f"{one(start)} to {one(end)}"


def export_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    # openpyxl gives every new workbook a default "Sheet"; nuke it so
    # the file only contains the sheets we explicitly add.
    wb.remove(wb.active)

    with session_scope() as sess:
        salon_row = sess.scalar(select(SalonRow).order_by(SalonRow.id).limit(1))
        location = salon_row.location if salon_row else ""

        # Location sheet
        ws = wb.create_sheet("Location")
        ws.append(["Locations", None])
        ws.append(["Location 1", location])

        # Hours sheet
        ws = wb.create_sheet("Hours")
        ws.append(["Day", "Open", "Close"])
        hours = {
            h.weekday: (h.open, h.close)
            for h in sess.scalars(select(HoursRow))
        }
        for day in WEEKDAYS:
            slot = hours.get(day)
            if not slot or slot[0] is None or slot[1] is None:
                ws.append([day, "NA", "NA"])
            else:
                ws.append([day, _fmt_hour_sheet(slot[0]), _fmt_hour_sheet(slot[1])])

        # Services sheet
        ws = wb.create_sheet("Services")
        ws.append(["name", "duration_minutes", "price"])
        for svc in sess.scalars(select(ServiceRow).order_by(ServiceRow.name)):
            ws.append([svc.name, svc.duration_minutes, svc.price])

        # Staff — Associates + Schedule
        staff_stmt = (
            select(StaffRow)
            .options(selectinload(StaffRow.services), selectinload(StaffRow.hours))
            .order_by(StaffRow.name)
        )
        staff_rows = sess.scalars(staff_stmt).all()

        max_services = max((len(s.services) for s in staff_rows), default=0)
        max_services = max(max_services, 4)
        ws = wb.create_sheet("Associates")
        ws.append(["Name"] + [f"Service {i+1}" for i in range(max_services)])
        for s in staff_rows:
            services = [svc.name for svc in s.services]
            row = [s.name] + [
                (services[i].title() if i < len(services) else None)
                for i in range(max_services)
            ]
            ws.append(row)

        ws = wb.create_sheet("Schedule")
        ws.append([None] + WEEKDAYS)
        for s in staff_rows:
            sched = {h.weekday: (h.start, h.end) for h in s.hours}
            row = [s.name]
            for day in WEEKDAYS:
                slot = sched.get(day)
                row.append(_fmt_range(*slot) if slot else None)
            ws.append(row)

        # Clients (Phase 2). Emitted before Appointments so the row order
        # in the workbook matches the FK dependency order — makes manual
        # inspection easier for anyone reading the file top to bottom.
        ws = wb.create_sheet("Clients")
        client_columns = [
            "id", "first_name", "last_name", "phone", "email",
            "gender", "notes", "created_at", "updated_at",
        ]
        ws.append(client_columns)
        for c in sess.scalars(select(ClientRow).order_by(ClientRow.id)):
            ws.append([
                c.id,
                c.first_name,
                c.last_name,
                c.phone,
                c.email,
                c.gender,
                c.notes,
                c.created_at.isoformat(timespec="seconds"),
                c.updated_at.isoformat(timespec="seconds"),
            ])

        # Appointments
        ws = wb.create_sheet("Appointments")
        columns = [
            "id", "created_at", "customer_name", "customer_phone",
            "stylist", "service", "date", "start_time", "end_time",
            "session_id", "client_id",
        ]
        ws.append(columns)
        appt_stmt = (
            select(Appointment)
            .options(selectinload(Appointment.staff), selectinload(Appointment.service))
            .order_by(Appointment.date, Appointment.start_time)
        )
        for a in sess.scalars(appt_stmt):
            ws.append([
                a.id,
                a.created_at.isoformat(timespec="seconds"),
                a.customer_name,
                a.customer_phone,
                a.staff.name,
                a.service.name,
                a.date.isoformat(),
                a.start_time.strftime("%H:%M"),
                a.end_time.strftime("%H:%M"),
                a.session_id,
                a.client_id,
            ])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.success(f"Exported to {path}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Export the salon DB back into a .xlsx workbook."
    )
    parser.add_argument("path", type=Path, help="Output workbook path")
    args = parser.parse_args()
    export_workbook(args.path)


if __name__ == "__main__":
    main()
